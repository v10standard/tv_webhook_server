from fastapi import FastAPI, Request
from openpyxl import load_workbook
from datetime import datetime
import os

app = FastAPI()

EXCEL_FILE = "Excel_Trade_Command_Center_Frame.xlsx"
RAW_DATA_SHEET = "Raw_Data"

@app.post("/tv-webhook")
async def receive_tv_alert(request: Request):
    payload = await request.json()

    symbol = payload.get("symbol", "N/A")
    tf = payload.get("tf", "N/A")
    time = payload.get("time", "N/A")
    rsi = payload.get("RSI", None)
    macd = payload.get("MACD", None)
    macd_signal = payload.get("MACD_Signal", None)
    wt1 = payload.get("WT1", None)
    wt2 = payload.get("WT2", None)

    if not os.path.exists(EXCEL_FILE):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = RAW_DATA_SHEET
        ws.append(["Timestamp", "Symbol", "Timeframe", "RSI", "MACD", "MACD_Signal", "WT1", "WT2"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    if RAW_DATA_SHEET not in wb.sheetnames:
        wb.create_sheet(RAW_DATA_SHEET)
    ws = wb[RAW_DATA_SHEET]

    if ws.max_row < 2:
        ws.append(["Timestamp", "Symbol", "Timeframe", "RSI", "MACD", "MACD_Signal", "WT1", "WT2"])

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        symbol, tf, rsi, macd, macd_signal, wt1, wt2
    ])

    wb.save(EXCEL_FILE)
    return {"status": "received"}